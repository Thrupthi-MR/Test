import pandas as pd

from io import BytesIO
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font


def read_excel_data(uploaded_file):
    """Read and clean excel file."""

    df = pd.read_excel(
        uploaded_file,
        sheet_name=0,
        header=7
    )

    df.columns = (
        df.columns
        .astype(str)
        .str.strip()
    )

    df = df.dropna(
        how="all"
    )

    return df


def create_summary_sheet(
    workbook,
    total_records,
    irregular_records
):
    """Create summary sheet."""

    summary = workbook.create_sheet(
        "Summary"
    )

    summary["A1"] = "Irregular Failure Report"
    summary["A2"] = "Summary"

    summary["A4"] = "Metric"
    summary["B4"] = "Value"

    summary["A5"] = "Total Records"
    summary["B5"] = total_records

    summary["A6"] = "Irregular Failures"
    summary["B6"] = irregular_records

    summary["A7"] = "Generated At"

    summary["B7"] = datetime.now().strftime(
        "%d-%m-%Y %H:%M:%S"
    )


def auto_adjust_width(ws):
    """Adjust column width automatically."""

    for column_cells in ws.columns:

        max_length = 0

        for cell in column_cells:

            try:
                if cell.value is not None:
                    max_length = max(
                        max_length,
                        len(str(cell.value))
                    )

            except Exception:
                pass

        ws.column_dimensions[
            column_cells[0].column_letter
        ].width = min(
            max_length + 3,
            80
        )


def generate_report(uploaded_file):
    """Generate Excel report."""

    df = read_excel_data(
        uploaded_file
    )

    irregular_df = df[
        df["CATEGORY"] == "IRREGULAR_TEST_FAILURE"
    ]

    wb = Workbook()

    ws = wb.active
    ws.title = "Irregular Failures"

    header_font = Font(
        bold=True,
        size=12
    )

    data_font = Font(
        size=11
    )

    # Write headers
    for col_num, column_name in enumerate(
        irregular_df.columns,
        start=1
    ):
        cell = ws.cell(
            row=1,
            column=col_num,
            value=column_name
        )

        cell.font = header_font

    # Write data
    for row_num, row_data in enumerate(
        irregular_df.values,
        start=2
    ):
        for col_num, value in enumerate(
            row_data,
            start=1
        ):
            cell = ws.cell(
                row=row_num,
                column=col_num,
                value=value
            )

            cell.font = data_font

    auto_adjust_width(ws)

    create_summary_sheet(
        wb,
        len(df),
        len(irregular_df)
    )

    timestamp = datetime.now().strftime(
        "%d-%m-%Y_%H-%M-%S"
    )

    output_path = (
        r"C:\Users\40054006\OneDrive - LTTS"
        r"\IrregularFailureReports"
        rf"\Irregular_Test_Failures_{timestamp}.xlsx"
    )

    wb.save(output_path)

    output = BytesIO()

    wb.save(output)

    output.seek(0)

    return (
        output,
        output_path,
        len(irregular_df)
    )

